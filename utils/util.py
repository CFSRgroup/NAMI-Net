import os
import random
import numpy as np
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def set_seed(seed):
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def create_excel_file(filepath):
    if not os.path.exists(filepath):
        workbook = Workbook()
        sheet = workbook.active
        headers = ['被试', 'ACC', '宏 F1', 'label 0 ACC', 'label 0 F1', 'label 1 ACC', 'label 1 F1', 'label 2 ACC', 'label 2 F1']
        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = 20
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(filepath)


def save_results_to_excel(filepath, subject_id, accuracy, label_acc, f1_scores, macro_f1):
    if os.path.exists(filepath):
        workbook = load_workbook(filepath)
    else:
        workbook = Workbook()

    sheet = workbook.active
    result_row = [f'被试{subject_id}', round(accuracy, 4), round(macro_f1, 4)]

    for acc, f1 in zip(label_acc, f1_scores):
        result_row.extend([round(float(acc.item()), 4), round(float(f1.item()), 4)])

    sheet.append(result_row)

    for cell in sheet[sheet.max_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filepath)

def save_overall_averages(filepath, subject_averages):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    avg_acc = np.mean([avg[0] for avg in subject_averages])
    std_acc = np.std([avg[0] for avg in subject_averages], ddof=1)
    avg_macro_f1 = np.mean([avg[1] for avg in subject_averages])
    std_macro_f1 = np.std([avg[1] for avg in subject_averages], ddof=1)

    num_labels = 3
    label_acc = []
    f1_scores = []

    for i in range(num_labels):
        label_values = [avg[i + 2] for avg in subject_averages]
        avg_label_acc = np.mean(label_values)
        std_label_acc = np.std(label_values, ddof=1)
        label_acc.append(f'{round(avg_label_acc, 4)}±{round(std_label_acc, 4)}')

        f1_values = [avg[i + 5] for avg in subject_averages]
        avg_f1_score = np.mean(f1_values)
        std_f1_score = np.std(f1_values, ddof=1)
        f1_scores.append(f'{round(avg_f1_score, 4)}±{round(std_f1_score, 4)}')

    overall_row = [
        '均值±标准差',
        f'{round(avg_acc, 4)}±{round(std_acc, 4)}',  # accuracy
        f'{round(avg_macro_f1, 4)}±{round(std_macro_f1, 4)}',  # macro F1
    ]


    for i in range(num_labels):
        overall_row.append(label_acc[i])  # label accuracy
        overall_row.append(f1_scores[i])  # f1 scores

    sheet.append(overall_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filepath)
