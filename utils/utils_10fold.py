import os
import random
import sys
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import scipy.io as sio
from sklearn.metrics import confusion_matrix, f1_score
from sklearn.model_selection import KFold
from torch.utils.data import DataLoader, TensorDataset
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def set_seed(seed):
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False

def create_excel_file(filepath):
    if not os.path.exists(filepath):
        workbook = Workbook()
        sheet = workbook.active
        headers = ['被试', '折数', 'ACC', '宏 F1', 'label 0 ACC', 'label 0 F1', 'label 1 ACC', 'label 1 F1', 'label 2 ACC', 'label 2 F1']
        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = 18
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(filepath)

def save_results_to_excel(filepath, subject_id, fold_id, accuracy, label_acc, f1_scores, macro_f1):
    if os.path.exists(filepath):
        workbook = load_workbook(filepath)
    else:
        workbook = Workbook()

    sheet = workbook.active
    result_row = [f'被试{subject_id}', f'fold{fold_id}', round(accuracy, 4), round(macro_f1, 4)]

    for acc, f1 in zip(label_acc, f1_scores):
        result_row.extend([round(float(acc), 4), round(float(f1), 4)])

    sheet.append(result_row)

    for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=3, max_col=10):
        for cell in row:
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filepath)

def save_fold_averages(filepath, subject_id, all_accs, all_macro_f1, all_label_acc, all_f1_scores):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    avg_acc = np.mean(all_accs)
    avg_macro_f1 = np.mean(all_macro_f1)
    avg_label_acc = np.mean(all_label_acc, axis=0)
    avg_f1_scores = np.mean(all_f1_scores, axis=0)

    avg_row = [f'被试{subject_id}', '平均值', avg_acc, avg_macro_f1]
    for acc, f1 in zip(avg_label_acc, avg_f1_scores):
        avg_row.extend([acc, f1])

    sheet.append(avg_row)

    new_row = sheet.max_row

    for cell in sheet[new_row][2:]:
        cell.number_format = '0.0000'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filepath)



def save_overall_averages(filepath, subject_averages):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    avg_acc = np.mean([avg[0] for avg in subject_averages])
    std_acc = np.std([avg[0] for avg in subject_averages], ddof=1)

    avg_macro_f1 = np.mean([avg[1] for avg in subject_averages])
    std_macro_f1 = np.std([avg[1] for avg in subject_averages], ddof=1)

    avg_label_acc = np.mean([avg[2] for avg in subject_averages], axis=0)
    std_label_acc = np.std([avg[2] for avg in subject_averages], axis=0, ddof=1)

    avg_f1_scores = np.mean([avg[3] for avg in subject_averages], axis=0)
    std_f1_scores = np.std([avg[3] for avg in subject_averages], axis=0, ddof=1)

    overall_row = ['总体平均值±标准差', '',
                   f"{avg_acc:.4f}±{std_acc:.4f}",
                   f"{avg_macro_f1:.4f}±{std_macro_f1:.4f}"]

    for acc, std_acc, f1, std_f1 in zip(avg_label_acc, std_label_acc, avg_f1_scores, std_f1_scores):
        overall_row.extend([f"{acc:.4f}±{std_acc:.4f}",
                            f"{f1:.4f}±{std_f1:.4f}"])

    sheet.append(overall_row)

    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(filepath)
